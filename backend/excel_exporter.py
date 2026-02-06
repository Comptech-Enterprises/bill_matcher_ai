import os
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Export bill matching results to Excel"""
    
    def __init__(self):
        # Define styles
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        self.profit_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.neutral_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.currency_format = '₹#,##0.00'
        self.percentage_format = '0.00%'
    
    def export_results(self, matched_results: Dict, summary: Dict, output_path: str) -> str:
        """
        Export matched results to Excel file
        
        Args:
            matched_results: Dictionary containing matched, unmatched_purchases, unmatched_sales
            summary: Summary statistics dictionary
            output_path: Path for the output Excel file
        
        Returns:
            Path to the generated Excel file
        """
        wb = Workbook()
        
        # Create sheets
        self._create_summary_sheet(wb, summary)
        self._create_matched_items_sheet(wb, matched_results['matched'])
        self._create_unmatched_purchases_sheet(wb, matched_results['unmatched_purchases'])
        self._create_unmatched_sales_sheet(wb, matched_results['unmatched_sales'])
        
        # Remove default sheet if it exists and is empty
        if 'Sheet' in wb.sheetnames and wb['Sheet'].max_row == 1:
            del wb['Sheet']
        
        # Save workbook
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        wb.save(output_path)
        
        return output_path
    
    def _create_summary_sheet(self, wb: Workbook, summary: Dict):
        """Create summary sheet with statistics"""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "Bill Matching Summary Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Summary data
        summary_data = [
            ("Metric", "Value"),
            ("Total Matched Items", summary.get('total_matched_items', 0)),
            ("Total Unmatched Purchases", summary.get('total_unmatched_purchases', 0)),
            ("Total Unmatched Sales", summary.get('total_unmatched_sales', 0)),
            ("", ""),
            ("Total Purchase Value", summary.get('total_purchase_value', 0)),
            ("Total Sale Value", summary.get('total_sale_value', 0)),
            ("Total Profit/Loss", summary.get('total_profit_loss', 0)),
            ("Profit/Loss Percentage", summary.get('total_profit_loss_percentage', 0)),
            ("", ""),
            ("Items with Profit", summary.get('profit_items_count', 0)),
            ("Items with Loss", summary.get('loss_items_count', 0)),
            ("", ""),
            ("Unmatched Purchase Value", summary.get('total_unmatched_purchase_value', 0)),
            ("Unmatched Sale Value", summary.get('total_unmatched_sale_value', 0)),
        ]
        
        start_row = 3
        for row_idx, (metric, value) in enumerate(summary_data, start=start_row):
            ws.cell(row=row_idx, column=1, value=metric)
            
            if row_idx == start_row:  # Header row
                ws.cell(row=row_idx, column=1).font = self.header_font
                ws.cell(row=row_idx, column=1).fill = self.header_fill
                ws.cell(row=row_idx, column=2).font = self.header_font
                ws.cell(row=row_idx, column=2).fill = self.header_fill
            
            cell = ws.cell(row=row_idx, column=2, value=value)
            
            # Apply currency format for monetary values
            if 'Value' in metric or 'Profit/Loss' in metric and 'Percentage' not in metric:
                cell.number_format = self.currency_format
            elif 'Percentage' in metric:
                cell.value = value / 100 if value else 0
                cell.number_format = self.percentage_format
            
            # Color code profit/loss
            if 'Profit/Loss' in metric and 'Percentage' not in metric:
                if value > 0:
                    cell.fill = self.profit_fill
                elif value < 0:
                    cell.fill = self.loss_fill
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_matched_items_sheet(self, wb: Workbook, matched_items: List[Dict]):
        """Create sheet for matched items"""
        ws = wb.create_sheet("Matched Items")

        headers = [
            "S.No", "Serial Number", "Item Name", "HSN Code", "Quantity",
            "Purchase Price", "Sale Price", "Profit/Loss", "Profit/Loss %"
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Write data
        for row_idx, item in enumerate(matched_items, start=2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = self.border
            ws.cell(row=row_idx, column=2, value=item.get('serial_number', 'N/A')).border = self.border
            ws.cell(row=row_idx, column=3, value=item.get('item_name', 'Unknown')).border = self.border
            ws.cell(row=row_idx, column=4, value=item.get('hsn_code', 'N/A')).border = self.border
            ws.cell(row=row_idx, column=5, value=item.get('quantity', 1)).border = self.border

            purchase_cell = ws.cell(row=row_idx, column=6, value=item.get('purchase_price', 0))
            purchase_cell.number_format = self.currency_format
            purchase_cell.border = self.border

            sale_cell = ws.cell(row=row_idx, column=7, value=item.get('sale_price', 0))
            sale_cell.number_format = self.currency_format
            sale_cell.border = self.border

            profit_loss = item.get('profit_loss', 0)
            pl_cell = ws.cell(row=row_idx, column=8, value=profit_loss)
            pl_cell.number_format = self.currency_format
            pl_cell.border = self.border

            pl_pct = item.get('profit_loss_percentage', 0) / 100
            pl_pct_cell = ws.cell(row=row_idx, column=9, value=pl_pct)
            pl_pct_cell.number_format = self.percentage_format
            pl_pct_cell.border = self.border

            # Color code based on profit/loss
            if profit_loss > 0:
                pl_cell.fill = self.profit_fill
                pl_pct_cell.fill = self.profit_fill
            elif profit_loss < 0:
                pl_cell.fill = self.loss_fill
                pl_pct_cell.fill = self.loss_fill
        
        # Adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_unmatched_purchases_sheet(self, wb: Workbook, unmatched_purchases: List[Dict]):
        """Create sheet for unmatched purchases"""
        ws = wb.create_sheet("Unmatched Purchases")

        headers = ["S.No", "Serial Number", "Item Name", "HSN Code", "Quantity", "Purchase Price"]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Write data
        for row_idx, item in enumerate(unmatched_purchases, start=2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = self.border
            ws.cell(row=row_idx, column=2, value=item.get('serial_number', 'N/A')).border = self.border
            ws.cell(row=row_idx, column=3, value=item.get('item_name', 'Unknown')).border = self.border
            ws.cell(row=row_idx, column=4, value=item.get('hsn_code', 'N/A')).border = self.border
            ws.cell(row=row_idx, column=5, value=item.get('quantity', 1)).border = self.border

            price_cell = ws.cell(row=row_idx, column=6, value=item.get('purchase_price', 0))
            price_cell.number_format = self.currency_format
            price_cell.border = self.border
            price_cell.fill = self.neutral_fill

        # Adjust column widths
        self._auto_adjust_columns(ws)

    def _create_unmatched_sales_sheet(self, wb: Workbook, unmatched_sales: List[Dict]):
        """Create sheet for unmatched sales"""
        ws = wb.create_sheet("Unmatched Sales")

        headers = ["S.No", "Serial Number", "Item Name", "HSN Code", "Quantity", "Sale Price"]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Write data
        for row_idx, item in enumerate(unmatched_sales, start=2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = self.border
            ws.cell(row=row_idx, column=2, value=item.get('serial_number', 'N/A')).border = self.border
            ws.cell(row=row_idx, column=3, value=item.get('item_name', 'Unknown')).border = self.border
            ws.cell(row=row_idx, column=4, value=item.get('hsn_code', 'N/A')).border = self.border
            ws.cell(row=row_idx, column=5, value=item.get('quantity', 1)).border = self.border

            price_cell = ws.cell(row=row_idx, column=6, value=item.get('sale_price', 0))
            price_cell.number_format = self.currency_format
            price_cell.border = self.border
            price_cell.fill = self.neutral_fill
        
        # Adjust column widths
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column].width = adjusted_width
